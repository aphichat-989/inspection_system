from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.http import Http404, HttpResponse, HttpResponseBadRequest, HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import TestSessionForm
from .models import DefectType, InspectionRound, InspectionSession, InspectionTest, Inspector
from .views_permissions import StaffRequiredMixin
from .views_helpers import (
    AUTO_COMPLETED_MESSAGE,
    build_next_session_number,
    get_round_result_types,
    record_inspection_round,
    test_summary,
)


class TestSessionListQuerysetMixin:
    def get_filter_values(self):
        return {
            "line_name": self.request.GET.get("line_name", "").strip(),
            "start_date": self.request.GET.get("start_date", "").strip(),
            "end_date": self.request.GET.get("end_date", "").strip(),
        }

    def get_base_queryset(self):
        return InspectionSession.objects.select_related("line", "test_condition", "inspector")

    def apply_filters(self, queryset):
        filters = self.get_filter_values()
        if filters["line_name"]:
            queryset = queryset.filter(line__name__icontains=filters["line_name"])
        if filters["start_date"]:
            queryset = queryset.filter(inspection_date__gte=filters["start_date"])
        if filters["end_date"]:
            queryset = queryset.filter(inspection_date__lte=filters["end_date"])
        return queryset

    def get_filtered_queryset(self):
        return self.apply_filters(self.get_base_queryset()).order_by("-inspection_date", "-created_at")


class InspectionListView(LoginRequiredMixin, TestSessionListQuerysetMixin, ListView):
    model = InspectionSession
    template_name = "inspection/list.html"
    context_object_name = "sessions"
    paginate_by = 15

    def get_queryset(self):
        return self.get_filtered_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.get_filter_values())
        return context


class InspectionSessionExportView(StaffRequiredMixin, TestSessionListQuerysetMixin, View):
    headers = [
        "Session Number",
        "Inspection Date",
        "Production Line",
        "Inspector",
        "Test Condition",
        "Status",
        "Number of Defect Tests",
        "Completed Tests",
        "Overall Comment",
        "Created At",
    ]
    report_table_headers = ["Defect", "Round", "Planned Round", "Result", "Inspection Time", "Comment"]

    def get_queryset(self):
        return self.get_filtered_queryset().annotate(
            defect_test_count=Count("tests", distinct=True),
            completed_test_count=Count(
                "tests",
                filter=Q(tests__status=InspectionTest.STATUS_FINISHED),
                distinct=True,
            ),
        )

    def get(self, request, *args, **kwargs):
        session_id = request.GET.get("session_id")
        if session_id:
            session = self.get_report_session(session_id)
            workbook = self.build_report_workbook(session)
            filename = f"inspection_report_{session.session_number}_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return self.build_response(workbook, filename)

        workbook = self.build_sessions_workbook()
        filename = f"inspection_sessions_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return self.build_response(workbook, filename)

    def get_report_session(self, session_id):
        try:
            return InspectionSession.objects.select_related("line", "test_condition", "inspector").prefetch_related(
                Prefetch(
                    "tests",
                    queryset=InspectionTest.objects.select_related("defect_type").prefetch_related(
                        Prefetch(
                            "rounds",
                            queryset=InspectionRound.objects.select_related("result_type").order_by("round_number", "created_at"),
                        )
                    ),
                )
            ).get(pk=session_id)
        except (InspectionSession.DoesNotExist, ValueError):
            raise Http404("Inspection session was not found")

    def build_sessions_workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Sessions"
        worksheet.append(self.headers)
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for session in self.get_queryset():
            created_at = timezone.localtime(session.created_at) if session.created_at else None
            worksheet.append(
                [
                    session.session_number,
                    session.inspection_date,
                    session.line.name if session.line_id else "-",
                    session.inspector.name if session.inspector_id else "-",
                    session.test_condition.name if session.test_condition_id else "-",
                    session.get_status_display(),
                    session.defect_test_count,
                    session.completed_test_count,
                    session.overall_comment,
                    created_at.replace(tzinfo=None) if created_at else None,
                ]
            )

        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].number_format = "yyyy-mm-dd"
        for row in worksheet.iter_rows(min_row=2, min_col=10, max_col=10):
            row[0].number_format = "yyyy-mm-dd hh:mm:ss"
        self.autosize_columns(worksheet)
        return workbook

    def build_report_workbook(self, session):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inspection Report"

        title_row = 1
        worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=7)
        title_cell = worksheet.cell(row=title_row, column=1, value="Inspection Session Summary")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        tests = list(session.tests.all())
        summaries = [test_summary(test) for test in tests]

        session_rows = [
            ("Session Number", self.safe_value(session.session_number)),
            ("Date", session.inspection_date.strftime("%d/%m/%Y") if session.inspection_date else "-"),
            ("Line", self.safe_value(session.line.name if session.line_id else None)),
            ("Inspection Type", self.safe_value(session.test_condition.name if session.test_condition_id else None)),
            ("Inspector", self.safe_value(session.inspector.name if session.inspector_id else None)),
        ]
        current_row = 3
        for label, value in session_rows:
            self.write_key_value_row(worksheet, current_row, label, value)
            current_row += 1

        current_row += 1
        self.write_section_header(worksheet, current_row, "Defect Summary", 7)
        summary_header_row = current_row + 1
        summary_headers = ["Defect", "Planned Rounds", "Completed Rounds", "FOUND", "NOT_FOUND", "Status", "Reason"]
        self.write_table_headers(worksheet, summary_header_row, summary_headers)
        current_row = summary_header_row + 1
        for summary in summaries:
            worksheet.append(
                [
                    self.get_test_display_name(summary["test"]),
                    summary["total_rounds"],
                    summary["completed"],
                    summary["found"],
                    summary["not_found"],
                    "Finished" if summary["is_finished"] else "In Progress",
                    self.safe_value(summary["stop_reason"]),
                ]
            )
            current_row += 1
        summary_last_row = max(current_row - 1, summary_header_row)
        self.style_table(worksheet, summary_header_row, summary_last_row, len(summary_headers))

        current_row += 1
        self.write_section_header(worksheet, current_row, "Overall Comment", 7)
        self.write_key_value_row(worksheet, current_row + 1, "Comment", self.safe_value(session.overall_comment))

        current_row += 3
        self.write_section_header(worksheet, current_row, "Round History", len(self.report_table_headers))
        table_header_row = current_row + 1
        self.write_table_headers(worksheet, table_header_row, self.report_table_headers)

        data_start_row = table_header_row + 1
        current_row = data_start_row
        for test in tests:
            defect_name = self.get_test_display_name(test)
            for round_item in test.rounds.all():
                inspection_time = timezone.localtime(round_item.created_at).replace(tzinfo=None) if round_item.created_at else None
                result = self.safe_value(round_item.result_type.name if round_item.result_type_id else None)
                worksheet.append(
                    [
                        defect_name,
                        f"Round {round_item.round_number}",
                        f"{round_item.round_number} / {test.total_rounds}",
                        result,
                        inspection_time,
                        self.safe_value(round_item.comment),
                    ]
                )
                current_row += 1

        last_data_row = max(current_row - 1, table_header_row)
        worksheet.freeze_panes = f"A{data_start_row}"
        last_column_letter = get_column_letter(len(self.report_table_headers))
        worksheet.auto_filter.ref = f"A{table_header_row}:{last_column_letter}{last_data_row}"
        self.style_report_table(worksheet, table_header_row, last_data_row)
        self.autosize_columns(worksheet)
        self.set_report_column_widths(worksheet)
        return workbook

    def write_section_header(self, worksheet, row, title, end_column):
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
        cell = worksheet.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="left")

    def write_key_value_row(self, worksheet, row, label, value):
        label_cell = worksheet.cell(row=row, column=1, value=label)
        value_cell = worksheet.cell(row=row, column=2, value=value)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left")
        value_cell.alignment = Alignment(horizontal="left", wrap_text=True)

    def write_table_headers(self, worksheet, row, headers):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row, column=column_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_table(self, worksheet, header_row, last_row, last_column):
        border_side = Side(style="thin", color="D9E2EC")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        for row in worksheet.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=last_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=2, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    def style_report_table(self, worksheet, table_header_row, last_data_row):
        last_column = len(self.report_table_headers)
        dark_fill = PatternFill(fill_type="solid", fgColor="000000")
        dark_border_side = Side(style="thin", color="1F1F1F")
        dark_border = Border(
            left=dark_border_side,
            right=dark_border_side,
            top=dark_border_side,
            bottom=dark_border_side,
        )
        header_font = Font(bold=True, color="FFFFFF")
        body_font = Font(color="FFFFFF")

        for row in worksheet.iter_rows(min_row=table_header_row, max_row=last_data_row, min_col=1, max_col=last_column):
            for cell in row:
                cell.fill = dark_fill
                cell.border = dark_border
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for cell in worksheet[table_header_row][:last_column]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=table_header_row + 1, max_row=last_data_row, min_col=2, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=table_header_row + 1, max_row=last_data_row, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=table_header_row + 1, max_row=last_data_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row in worksheet.iter_rows(min_row=table_header_row + 1, max_row=last_data_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row_index in range(table_header_row, last_data_row + 1):
            worksheet.row_dimensions[row_index].height = 28

    def set_report_column_widths(self, worksheet):
        minimum_widths = {
            "A": 24,
            "B": 14,
            "C": 18,
            "D": 18,
            "E": 28,
            "F": 28,
        }
        for column_letter, width in minimum_widths.items():
            worksheet.column_dimensions[column_letter].width = max(
                worksheet.column_dimensions[column_letter].width or 0,
                width,
            )

    def autosize_columns(self, worksheet):
        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    def format_result(self, value):
        normalized = self.safe_value(value)
        if normalized == "-":
            return "-"
        return normalized.upper().replace(" ", "_")

    def safe_value(self, value):
        if value is None:
            return "-"
        value = str(value).strip()
        return value or "-"

    def get_test_display_name(self, test):
        return self.safe_value(test.defect_type.name if test.defect_type_id else test.test_name)

    def build_response(self, workbook, filename):
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

def get_logged_in_inspector(user):
    name = user.get_full_name() or user.username
    inspector, _ = Inspector.objects.update_or_create(
        name=name,
        defaults={"is_active": True},
    )
    return inspector

class TestSessionContextMixin:
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_defect_context(self, form=None):
        active_defects = DefectType.objects.filter(is_active=True).order_by("name")
        selected_ids = set(self.request.POST.getlist("defects"))
        round_values = {}
        if getattr(self, "object", None) and self.object.pk and not selected_ids:
            tests = self.object.tests.select_related("defect_type")
            selected_ids = {str(test.defect_type_id) for test in tests if test.defect_type_id}
            round_values = {str(test.defect_type_id): test.total_rounds for test in tests if test.defect_type_id}
        else:
            for defect in active_defects:
                round_values[str(defect.pk)] = self.request.POST.get(f"rounds_{defect.pk}", 1)
        defect_cards = []
        for defect in active_defects:
            defect_id = str(defect.pk)
            defect_cards.append(
                {
                    "defect": defect,
                    "selected": defect_id in selected_ids,
                    "rounds": round_values.get(defect_id, self.request.POST.get(f"rounds_{defect.pk}", 1)),
                }
            )
        return {
            "defect_cards": defect_cards,
            "has_active_inspectors": Inspector.objects.filter(is_active=True).exists(),
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self.get_defect_context(context.get("form")))
        return context


class InspectionCreateView(LoginRequiredMixin, TestSessionContextMixin, CreateView):
    model = InspectionSession
    form_class = TestSessionForm
    template_name = "inspection/form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["session_number"] = build_next_session_number()
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        selected_defects = DefectType.objects.filter(pk__in=self.request.POST.getlist("defects"), is_active=True)
        if not selected_defects.exists():
            form.add_error(None, "Please select at least one defect to test.")
            return self.form_invalid(form)

        self.object = form.save(commit=False)
        self.object.inspector = get_logged_in_inspector(self.request.user)
        self.object.session_number = build_next_session_number()
        self.object.status = InspectionSession.STATUS_IN_PROGRESS
        self.object.save()
        for defect in selected_defects:
            total_rounds = max(1, int(self.request.POST.get(f"rounds_{defect.pk}") or 1))
            InspectionTest.objects.create(
                session=self.object,
                defect_type=defect,
                test_name=str(defect.pk),
                total_rounds=total_rounds,
            )
        messages.success(self.request, "Test Session created. Start testing now.")
        return HttpResponseRedirect(reverse("inspection:detail", kwargs={"pk": self.object.pk}))


class InspectionUpdateView(LoginRequiredMixin, TestSessionContextMixin, UpdateView):
    model = InspectionSession
    form_class = TestSessionForm
    template_name = "inspection/form.html"
    success_url = reverse_lazy("inspection:list")

    def get_queryset(self):
        return InspectionSession.objects.prefetch_related("tests__rounds")

    @transaction.atomic
    def form_valid(self, form):
        form.instance.inspector = get_logged_in_inspector(self.request.user)
        self.object = form.save()
        selected_defects = DefectType.objects.filter(pk__in=self.request.POST.getlist("defects"), is_active=True)
        existing_tests = {test.defect_type_id: test for test in self.object.tests.select_related("defect_type") if test.defect_type_id}
        for defect in selected_defects:
            total_rounds = max(1, int(self.request.POST.get(f"rounds_{defect.pk}") or 1))
            test = existing_tests.get(defect.pk)
            if not test:
                test = InspectionTest.objects.create(session=self.object, defect_type=defect, test_name=str(defect.pk), total_rounds=total_rounds)
            else:
                test.total_rounds = total_rounds
                test.save(update_fields=["total_rounds"])

        messages.success(self.request, "Test Session updated.")
        return HttpResponseRedirect(reverse("inspection:detail", kwargs={"pk": self.object.pk}))


class InspectionDetailView(LoginRequiredMixin, DetailView):
    model = InspectionSession
    template_name = "inspection/detail.html"
    context_object_name = "session"

    def get_queryset(self):
        return InspectionSession.objects.select_related("line", "test_condition", "inspector").prefetch_related(
            Prefetch(
                "tests",
                queryset=InspectionTest.objects.select_related("defect_type").prefetch_related(
                    Prefetch("rounds", queryset=InspectionRound.objects.select_related("result_type").order_by("round_number"))
                ),
            )
        )

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        action = request.POST.get("action", "save")
        current_index = int(request.POST.get("current_index") or 0)
        current_test_id = request.POST.get("current_test_id")
        result_value = request.POST.get("result", "")
        show_finished_summary = False

        if result_value:
            if not current_test_id:
                return HttpResponseBadRequest("No inspection test was selected.")
            result, error = record_inspection_round(
                self.object,
                current_test_id,
                result_value,
                request.POST.get("comment", ""),
            )
            if error:
                return HttpResponseBadRequest(error)
            if result["auto_completed"]:
                messages.info(request, AUTO_COMPLETED_MESSAGE)
                show_finished_summary = True
            else:
                messages.success(request, "Inspection round recorded.")
                show_finished_summary = result["test"].status == InspectionTest.STATUS_FINISHED
        elif action == "save":
            return HttpResponseBadRequest("Please select FOUND or NOT_FOUND before submitting.")

        tests_count = self.object.tests.count()
        if action == "finish":
            self.object.status = InspectionSession.STATUS_COMPLETED
            self.object.save(update_fields=["status"])
            messages.success(request, "Test Session finished.")
            return HttpResponseRedirect(f"{reverse('inspection:detail', kwargs={'pk': self.object.pk})}?summary=1")
        if show_finished_summary:
            return HttpResponseRedirect(
                f"{reverse('inspection:detail', kwargs={'pk': self.object.pk})}?defect={current_index}&summary=1"
            )
        if action == "previous":
            current_index = max(0, current_index - 1)
        elif action == "next":
            current_index = min(max(tests_count - 1, 0), current_index + 1)
        return HttpResponseRedirect(f"{reverse('inspection:detail', kwargs={'pk': self.object.pk})}?defect={current_index}")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        found_result, not_found_result = get_round_result_types()
        tests = list(self.object.tests.all())
        current_index = int(self.request.GET.get("defect") or 0)
        current_index = min(max(current_index, 0), max(len(tests) - 1, 0)) if tests else 0
        summaries = [test_summary(test) for test in tests]
        current_summary = summaries[current_index] if summaries else None
        current_test = tests[current_index] if tests else None
        current_round_number = current_summary["next_round_number"] if current_summary else 0
        round_progress_percent = (
            round((current_summary["completed"] / current_summary["total_rounds"]) * 100)
            if current_summary and current_summary["total_rounds"]
            else 0
        )
        show_summary = self.request.GET.get("summary") == "1" or (
            current_summary and current_summary["is_finished"]
        )
        context.update(
            {
                "tests": tests,
                "current_test": current_test,
                "current_index": current_index,
                "display_index": current_index + 1,
                "total_tests": len(tests),
                "progress_percent": round(((current_index + 1) / len(tests)) * 100) if tests else 0,
                "round_progress_percent": round_progress_percent,
                "summaries": summaries,
                "current_summary": current_summary,
                "found_result": found_result,
                "not_found_result": not_found_result,
                "current_round_number": current_round_number,
                "show_summary": show_summary,
            }
        )
        return context


class InspectionBulkDeleteView(StaffRequiredMixin, View):
    success_url = reverse_lazy("inspection:list")

    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist("selected_sessions")
        if not selected_ids:
            messages.warning(request, "Please select at least one Test Session to delete.")
            return HttpResponseRedirect(self.success_url)

        queryset = InspectionSession.objects.filter(pk__in=selected_ids)
        deleted_count = queryset.count()
        if not deleted_count:
            messages.warning(request, "Selected Test Sessions were not found.")
            return HttpResponseRedirect(self.success_url)

        queryset.delete()
        messages.success(request, f"Deleted {deleted_count} Test Session(s).")
        return HttpResponseRedirect(self.success_url)


class InspectionDeleteView(StaffRequiredMixin, DeleteView):
    model = InspectionSession
    template_name = "inspection/confirm_delete.html"
    success_url = reverse_lazy("inspection:list")

    def form_valid(self, form):
        messages.success(self.request, "Test Session deleted.")
        return super().form_valid(form)



