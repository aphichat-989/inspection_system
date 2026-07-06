from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import (
    DefectType,
    InspectionRound,
    InspectionSession,
    InspectionTest,
    Inspector,
    ProductionLine,
    TestCondition,
)
from .views import NOT_FOUND_STOP_REASON


class InspectionWorkflowTests(TestCase):
    def setUp(self):
        self.line = ProductionLine.objects.create(name="Line A")
        self.condition = TestCondition.objects.create(name="Normal Light")
        self.inspector = Inspector.objects.create(name="Inspector A")
        self.defect = DefectType.objects.create(name="Scratch")

    def create_session_with_test(self, total_rounds=30):
        session = InspectionSession.objects.create(
            session_number="TS20260703-0001",
            inspection_date=date(2026, 7, 3),
            line=self.line,
            test_condition=self.condition,
            inspector=self.inspector,
            status=InspectionSession.STATUS_IN_PROGRESS,
        )
        test = InspectionTest.objects.create(
            session=session,
            defect_type=self.defect,
            test_name=str(self.defect.pk),
            total_rounds=total_rounds,
        )
        return session, test

    def create_export_session(
        self,
        session_number,
        inspection_date=None,
        line=None,
        completed_tests=0,
        in_progress_tests=0,
        overall_comment="",
    ):
        session = InspectionSession.objects.create(
            session_number=session_number,
            inspection_date=inspection_date or date(2026, 7, 3),
            line=line or self.line,
            test_condition=self.condition,
            inspector=self.inspector,
            status=InspectionSession.STATUS_IN_PROGRESS,
            overall_comment=overall_comment,
        )
        for index in range(completed_tests):
            defect = DefectType.objects.create(name=f"{session_number} Completed {index}")
            InspectionTest.objects.create(
                session=session,
                defect_type=defect,
                test_name=str(defect.pk),
                total_rounds=1,
                completed_rounds=1,
                status=InspectionTest.STATUS_FINISHED,
            )
        for index in range(in_progress_tests):
            defect = DefectType.objects.create(name=f"{session_number} In Progress {index}")
            InspectionTest.objects.create(
                session=session,
                defect_type=defect,
                test_name=str(defect.pk),
                total_rounds=1,
                status=InspectionTest.STATUS_IN_PROGRESS,
            )
        return session

    def export_workbook(self, query_string=""):
        url = reverse("inspection:session_export")
        response = self.client.get(f"{url}?{query_string}" if query_string else url)
        workbook = load_workbook(BytesIO(response.content))
        return response, workbook.active

    def test_create_session_does_not_precreate_rounds(self):
        response = self.client.post(
            reverse("inspection:create"),
            {
                "inspection_date": "2026-07-03",
                "line": self.line.pk,
                "test_condition": self.condition.pk,
                "inspector": self.inspector.pk,
                "overall_comment": "",
                "defects": [self.defect.pk],
                f"rounds_{self.defect.pk}": "30",
            },
        )

        self.assertEqual(response.status_code, 302)
        test = InspectionTest.objects.get(defect_type=self.defect)
        self.assertEqual(test.total_rounds, 30)
        self.assertEqual(test.rounds.count(), 0)

    def test_detail_displays_only_current_round_without_creating_database_rows(self):
        session, test = self.create_session_with_test(total_rounds=30)

        response = self.client.get(reverse("inspection:detail", kwargs={"pk": session.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Current Round")
        self.assertContains(response, "1 / 30")
        self.assertNotContains(response, "Round 30")
        self.assertContains(response, "FOUND")
        self.assertContains(response, "NOT_FOUND")
        self.assertContains(response, "Note")
        self.assertContains(response, "Save")
        self.assertEqual(test.rounds.count(), 0)

    def test_saving_round_creates_one_record_and_advances_current_round(self):
        session, test = self.create_session_with_test(total_rounds=30)
        url = reverse("inspection:detail", kwargs={"pk": session.pk})

        response = self.client.post(
            url,
            {
                "action": "save",
                "current_index": "0",
                "current_test_id": str(test.pk),
                "result": "found",
                "comment": "ok",
            },
        )

        self.assertEqual(response.status_code, 302)
        test.refresh_from_db()
        self.assertEqual(test.completed_rounds, 1)
        self.assertEqual(test.rounds.count(), 1)
        round_item = test.rounds.get()
        self.assertEqual(round_item.round_number, 1)
        self.assertEqual(round_item.comment, "ok")

        response = self.client.get(url)
        self.assertContains(response, "2 / 30")
        self.assertEqual(test.rounds.count(), 1)

        response = self.client.get(f"{url}?summary=1")
        self.assertContains(response, "ประวัติรอบทดสอบ")
        self.assertContains(response, "Round 1")
        self.assertContains(response, "1 / 30")
        self.assertContains(response, "ok")

    def test_save_requires_result(self):
        session, test = self.create_session_with_test(total_rounds=30)
        response = self.client.post(
            reverse("inspection:detail", kwargs={"pk": session.pk}),
            {
                "action": "save",
                "current_index": "0",
                "current_test_id": str(test.pk),
                "comment": "missing result",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(test.rounds.count(), 0)

    def test_not_found_four_finishes_test_and_rejects_extra_rounds(self):
        session, test = self.create_session_with_test(total_rounds=30)
        url = reverse("inspection:detail", kwargs={"pk": session.pk})

        for index in range(4):
            response = self.client.post(
                url,
                {
                    "action": "save",
                    "current_index": "0",
                    "current_test_id": str(test.pk),
                    "result": "not_found",
                    "comment": "",
                },
            )
            self.assertEqual(response.status_code, 302)
            if index == 3:
                self.assertIn("summary=1", response["Location"])

        test.refresh_from_db()
        self.assertEqual(test.status, InspectionTest.STATUS_FINISHED)
        self.assertEqual(test.stop_reason, NOT_FOUND_STOP_REASON)
        self.assertEqual(test.completed_rounds, 4)
        self.assertEqual(test.rounds.count(), 4)

        response = self.client.post(
            url,
            {
                "action": "save",
                "current_index": "0",
                "current_test_id": str(test.pk),
                "result": "found",
                "comment": "manual extra request",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("already finished", response.content.decode())
        self.assertEqual(InspectionRound.objects.filter(inspection_test=test).count(), 4)

    def test_completing_planned_rounds_finishes_test_and_shows_summary(self):
        session, test = self.create_session_with_test(total_rounds=2)
        url = reverse("inspection:detail", kwargs={"pk": session.pk})

        for index in range(2):
            response = self.client.post(
                url,
                {
                    "action": "save",
                    "current_index": "0",
                    "current_test_id": str(test.pk),
                    "result": "found",
                    "comment": "",
                },
            )
            self.assertEqual(response.status_code, 302)
            if index == 1:
                self.assertIn("summary=1", response["Location"])

        test.refresh_from_db()
        session.refresh_from_db()
        self.assertEqual(test.status, InspectionTest.STATUS_FINISHED)
        self.assertEqual(test.completed_rounds, 2)
        self.assertEqual(test.rounds.count(), 2)
        self.assertEqual(session.status, InspectionSession.STATUS_COMPLETED)
    def test_session_list_keeps_existing_filter_and_navigation(self):
        matching_line = ProductionLine.objects.create(name="List Filter Line")
        other_line = ProductionLine.objects.create(name="List Other Line")
        matching_session = self.create_export_session("TS20260703-3001", line=matching_line)
        self.create_export_session("TS20260703-3002", line=other_line)

        response = self.client.get(reverse("inspection:list"), {"line_name": "Filter"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "TS20260703-3001")
        self.assertNotContains(response, "TS20260703-3002")
        self.assertContains(response, reverse("inspection:detail", kwargs={"pk": matching_session.pk}))
        self.assertContains(response, reverse("inspection:create"))
        self.assertContains(response, f"{reverse('inspection:session_export')}?line_name=Filter")

    def test_export_sessions_without_filters_includes_all_sessions(self):
        self.create_export_session("TS20260703-1001", completed_tests=1, in_progress_tests=1)
        self.create_export_session("TS20260704-1002", inspection_date=date(2026, 7, 4))

        response, sheet = self.export_workbook()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet["A1"].value, "Session Number")
        self.assertEqual(sheet["G2"].value, 0)
        self.assertEqual(sheet["H2"].value, 0)
        self.assertEqual(sheet["G3"].value, 2)
        self.assertEqual(sheet["H3"].value, 1)

    def test_export_sessions_with_date_filters(self):
        self.create_export_session("TS20260701-1001", inspection_date=date(2026, 7, 1))
        self.create_export_session("TS20260703-1002", inspection_date=date(2026, 7, 3))
        self.create_export_session("TS20260705-1003", inspection_date=date(2026, 7, 5))

        response, sheet = self.export_workbook("start_date=2026-07-02&end_date=2026-07-04")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "TS20260703-1002")

    def test_export_sessions_with_search_filter(self):
        matching_line = ProductionLine.objects.create(name="Search Line Alpha")
        other_line = ProductionLine.objects.create(name="Other Line")
        self.create_export_session("TS20260703-1001", line=matching_line)
        self.create_export_session("TS20260703-1002", line=other_line)

        response, sheet = self.export_workbook("line_name=Alpha")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["A2"].value, "TS20260703-1001")
        self.assertEqual(sheet["C2"].value, "Search Line Alpha")

    def test_export_sessions_across_multiple_pagination_pages(self):
        paginated_line = ProductionLine.objects.create(name="Paginated Export Line")
        for index in range(16):
            self.create_export_session(f"TS20260703-{2000 + index}", line=paginated_line)
        self.create_export_session("TS20260703-9999", line=ProductionLine.objects.create(name="Excluded Line"))

        response, sheet = self.export_workbook("line_name=Paginated")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet.max_row, 17)

    def test_export_sessions_empty_result_returns_header_only(self):
        self.create_export_session("TS20260703-1001")

        response, sheet = self.export_workbook("line_name=NoMatchingLine")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet.max_row, 1)
        self.assertEqual([cell.value for cell in sheet[1]], [
            "Session Number",
            "Inspection Date",
            "Production Line",
            "Inspector",
            "Test Condition",
            "Status",
            "Number of Defect Tests",
            "Completed Tests",
            "Overall Comment",
            "Created At",
        ])

    def test_export_sessions_response_headers_and_filename(self):
        self.create_export_session("TS20260703-1001")

        response, sheet = self.export_workbook()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertRegex(
            response["Content-Disposition"],
            r'^attachment; filename="inspection_sessions_\d{8}_\d{6}\.xlsx"$',
        )
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["B2"].number_format, "yyyy-mm-dd")

    def test_bulk_delete_removes_only_selected_sessions(self):
        selected_session, _ = self.create_session_with_test(total_rounds=5)
        kept_session = InspectionSession.objects.create(
            session_number="TS20260703-0002",
            inspection_date=date(2026, 7, 3),
            line=self.line,
            test_condition=self.condition,
            inspector=self.inspector,
            status=InspectionSession.STATUS_IN_PROGRESS,
        )

        response = self.client.post(
            reverse("inspection:bulk_delete"),
            {"selected_sessions": [str(selected_session.pk)]},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(InspectionSession.objects.filter(pk=selected_session.pk).exists())
        self.assertTrue(InspectionSession.objects.filter(pk=kept_session.pk).exists())
